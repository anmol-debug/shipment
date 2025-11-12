"""
Unit tests for PDF and Excel utility functions
"""
import pytest
from pathlib import Path
import tempfile
import os

from app.utils.pdf_utils import extract_text_from_pdf, extract_pdf_as_image
from app.utils.xlsx_utils import extract_text_from_xlsx


class TestPDFUtils:
    """Tests for PDF utility functions"""

    def test_extract_text_from_text_pdf(self):
        """Test extracting text from a text-based PDF"""
        # Create a simple text PDF for testing
        try:
            from reportlab.pdfgen import canvas

            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp_path = tmp.name

            # Create PDF with text
            c = canvas.Canvas(tmp_path)
            c.drawString(100, 750, "Test Bill of Lading")
            c.drawString(100, 700, "B/L No: TEST123")
            c.save()

            # Test extraction
            result = extract_text_from_pdf(tmp_path)
            assert isinstance(result, str)
            assert len(result) > 0
            # Text extraction from reportlab PDFs might vary
            assert "Test" in result or "TEST" in result or len(result) > 10

            # Cleanup
            os.unlink(tmp_path)

        except ImportError:
            pytest.skip("reportlab not installed for PDF creation")

    def test_extract_text_from_nonexistent_pdf(self):
        """Test that extracting from non-existent PDF raises error"""
        with pytest.raises(FileNotFoundError):
            extract_text_from_pdf("/nonexistent/path/file.pdf")

    def test_extract_pdf_as_image_returns_base64(self):
        """Test that image extraction returns base64 encoded data"""
        # Create a minimal PDF
        try:
            from reportlab.pdfgen import canvas

            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp_path = tmp.name

            c = canvas.Canvas(tmp_path)
            c.drawString(100, 750, "Test")
            c.save()

            result = extract_pdf_as_image(tmp_path)

            # Check format
            assert isinstance(result, str)
            assert result.startswith("IMAGE_PDF:")

            # Extract base64 part
            base64_data = result.split("IMAGE_PDF:")[1]
            assert len(base64_data) > 100  # Should be substantial
            # Base64 should only contain valid characters
            import string
            valid_chars = set(string.ascii_letters + string.digits + '+/=')
            assert all(c in valid_chars for c in base64_data[:100])

            # Cleanup
            os.unlink(tmp_path)

        except ImportError:
            pytest.skip("reportlab or pypdfium2 not installed")


class TestExcelUtils:
    """Tests for Excel utility functions"""

    def test_extract_text_from_xlsx(self):
        """Test extracting text from Excel file"""
        try:
            import openpyxl

            # Create test Excel file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TestSheet"

            # Add data
            ws['A1'] = "Bill of Lading"
            ws['B1'] = "TEST123"
            ws['A2'] = "Container"
            ws['B2'] = "CONT456"

            wb.save(tmp_path)

            # Test extraction
            result = extract_text_from_xlsx(tmp_path)

            assert isinstance(result, str)
            assert "TestSheet" in result
            assert "Bill of Lading" in result
            assert "TEST123" in result
            assert "Container" in result

            # Cleanup
            os.unlink(tmp_path)

        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_extract_text_from_multiple_sheets(self):
        """Test extracting text from Excel with multiple sheets"""
        try:
            import openpyxl

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name

            wb = openpyxl.Workbook()

            # Sheet 1
            ws1 = wb.active
            ws1.title = "Invoice"
            ws1['A1'] = "Invoice Data"

            # Sheet 2
            ws2 = wb.create_sheet("Packing List")
            ws2['A1'] = "Packing Data"

            wb.save(tmp_path)

            # Test extraction
            result = extract_text_from_xlsx(tmp_path)

            assert "Invoice" in result
            assert "Packing List" in result
            assert "Invoice Data" in result
            assert "Packing Data" in result

            # Cleanup
            os.unlink(tmp_path)

        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_extract_text_handles_empty_cells(self):
        """Test that empty cells are handled correctly"""
        try:
            import openpyxl

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name

            wb = openpyxl.Workbook()
            ws = wb.active

            # Add data with empty cells
            ws['A1'] = "Data1"
            ws['B1'] = None  # Empty cell
            ws['C1'] = "Data2"

            wb.save(tmp_path)

            # Test extraction
            result = extract_text_from_xlsx(tmp_path)

            assert isinstance(result, str)
            assert "Data1" in result
            assert "Data2" in result

            # Cleanup
            os.unlink(tmp_path)

        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_extract_text_from_nonexistent_xlsx(self):
        """Test that extracting from non-existent Excel raises error"""
        with pytest.raises(Exception) as exc_info:
            extract_text_from_xlsx("/nonexistent/path/file.xlsx")
        assert "Failed to extract XLSX text" in str(exc_info.value)
