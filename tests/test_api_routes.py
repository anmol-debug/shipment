"""
Unit tests for API routes
"""
import pytest
from fastapi.testclient import TestClient
from pathlib import Path
import json
import tempfile
import shutil

from main import app
from app.api.routes import UPLOAD_DIR, STORAGE_DIR

client = TestClient(app)


class TestHealthEndpoint:
    """Tests for health check endpoint"""

    def test_health_check_returns_ok(self):
        """Test that health endpoint returns 200 OK"""
        response = client.get("/api/health")
        assert response.status_code == 200
        assert response.json() == {"status": "ok", "message": "API is running"}


class TestSaveEndpoint:
    """Tests for save extraction endpoint"""

    def test_save_extraction_success(self):
        """Test successful save of extraction data"""
        test_data = {
            "extractedData": {
                "billOfLadingNumber": "TEST123",
                "containerNumber": "CONT456",
                "consigneeName": "Test Company",
                "consigneeAddress": "123 Test St",
                "dateOfExport": "01/01/2024",
                "lineItemsCount": 10,
                "averageGrossWeight": "100 KG",
                "averagePrice": "$50.00"
            },
            "files": []
        }

        response = client.post("/api/save", json=test_data)
        assert response.status_code == 200
        assert response.json()["success"] is True
        assert "TEST123" in response.json()["message"]

        # Cleanup
        test_folder = STORAGE_DIR / "TEST123"
        if test_folder.exists():
            shutil.rmtree(test_folder)

    def test_save_extraction_missing_bl_number(self):
        """Test save fails without B/L number"""
        test_data = {
            "extractedData": {
                "containerNumber": "CONT456"
            },
            "files": []
        }

        response = client.post("/api/save", json=test_data)
        assert response.status_code == 400
        assert "B/L number is required" in response.json()["detail"]

    def test_save_extraction_sanitizes_bl_number(self):
        """Test that B/L number is sanitized for folder name"""
        test_data = {
            "extractedData": {
                "billOfLadingNumber": "TEST/123-ABC",
                "containerNumber": "CONT456"
            },
            "files": []
        }

        response = client.post("/api/save", json=test_data)
        assert response.status_code == 200

        # Check sanitized folder exists
        sanitized_folder = STORAGE_DIR / "TEST_123_ABC"
        assert sanitized_folder.exists()

        # Cleanup
        if sanitized_folder.exists():
            shutil.rmtree(sanitized_folder)


class TestRetrieveEndpoint:
    """Tests for retrieve extraction endpoint"""

    def test_retrieve_existing_extraction(self):
        """Test retrieving an existing extraction"""
        # Setup: Create a test extraction
        bl_number = "RETRIEVE_TEST"
        test_folder = STORAGE_DIR / bl_number
        test_folder.mkdir(exist_ok=True)

        test_metadata = {
            "billOfLadingNumber": bl_number,
            "savedAt": "2024-01-01T00:00:00",
            "extractedData": {"test": "data"},
            "files": []
        }

        with open(test_folder / "metadata.json", "w") as f:
            json.dump(test_metadata, f)

        # Test retrieval
        response = client.get(f"/api/retrieve/{bl_number}")
        assert response.status_code == 200
        assert response.json()["success"] is True
        assert response.json()["data"]["billOfLadingNumber"] == bl_number

        # Cleanup
        shutil.rmtree(test_folder)

    def test_retrieve_nonexistent_extraction(self):
        """Test retrieving a non-existent extraction returns 404"""
        response = client.get("/api/retrieve/NONEXISTENT123")
        assert response.status_code == 404
        assert "No saved data found" in response.json()["detail"]


class TestListEndpoint:
    """Tests for list saved extractions endpoint"""

    def test_list_empty_extractions(self):
        """Test listing when no extractions exist"""
        # Temporarily rename storage dir to ensure empty
        temp_storage = STORAGE_DIR.parent / "temp_storage"
        if STORAGE_DIR.exists():
            STORAGE_DIR.rename(temp_storage)
        STORAGE_DIR.mkdir(exist_ok=True)

        response = client.get("/api/list")
        assert response.status_code == 200
        assert response.json()["success"] is True
        assert response.json()["count"] == 0

        # Restore
        shutil.rmtree(STORAGE_DIR)
        if temp_storage.exists():
            temp_storage.rename(STORAGE_DIR)

    def test_list_multiple_extractions(self):
        """Test listing multiple saved extractions"""
        # Setup: Create test extractions
        test_bls = ["LIST_TEST1", "LIST_TEST2"]

        for bl in test_bls:
            test_folder = STORAGE_DIR / bl
            test_folder.mkdir(exist_ok=True)

            test_metadata = {
                "billOfLadingNumber": bl,
                "savedAt": "2024-01-01T00:00:00",
                "extractedData": {
                    "containerNumber": f"CONT_{bl}",
                    "consigneeName": f"Company {bl}"
                }
            }

            with open(test_folder / "metadata.json", "w") as f:
                json.dump(test_metadata, f)

        # Test listing
        response = client.get("/api/list")
        assert response.status_code == 200
        assert response.json()["success"] is True
        assert response.json()["count"] >= 2

        # Check that our test items are in the list
        bl_numbers = [item["billOfLadingNumber"] for item in response.json()["data"]]
        for bl in test_bls:
            assert bl in bl_numbers

        # Cleanup
        for bl in test_bls:
            test_folder = STORAGE_DIR / bl
            if test_folder.exists():
                shutil.rmtree(test_folder)


class TestExcelPreviewEndpoint:
    """Tests for Excel preview endpoint"""

    def test_excel_preview_file_not_found(self):
        """Test Excel preview returns error for non-existent file"""
        response = client.get("/api/excel-preview/nonexistent.xlsx")
        # The endpoint returns 500 for errors (as per current implementation)
        assert response.status_code in [404, 500]
        if response.status_code == 404:
            assert "File not found" in response.json()["detail"]
        else:
            assert "Failed to generate preview" in response.json()["detail"]

    def test_excel_preview_returns_html(self):
        """Test Excel preview returns HTML content"""
        # This test requires an actual Excel file in uploads directory
        # Create a simple test Excel file
        try:
            import openpyxl

            test_file = UPLOAD_DIR / "test_preview.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "Header1"
            ws['B1'] = "Header2"
            ws['A2'] = "Data1"
            ws['B2'] = "Data2"
            wb.save(test_file)

            response = client.get("/api/excel-preview/test_preview.xlsx")
            assert response.status_code == 200
            assert "text/html" in response.headers["content-type"]
            assert "<table>" in response.text
            assert "Header1" in response.text

            # Cleanup
            test_file.unlink()
        except ImportError:
            pytest.skip("openpyxl not installed")
