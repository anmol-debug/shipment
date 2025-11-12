"""
Unit tests for document processor
"""
import pytest
from pathlib import Path
import tempfile
import os

from app.services.document_processor import process_documents


class TestDocumentProcessor:
    """Tests for document processing service"""

    def test_process_single_pdf(self):
        """Test processing a single PDF file"""
        try:
            from reportlab.pdfgen import canvas

            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp_path = tmp.name

            # Create PDF with text
            c = canvas.Canvas(tmp_path)
            c.drawString(100, 750, "Bill of Lading TEST123")
            c.save()

            # Process
            result = process_documents([tmp_path])

            assert isinstance(result, str)
            assert len(result) > 0

            # Cleanup
            os.unlink(tmp_path)

        except ImportError:
            pytest.skip("reportlab not installed")

    def test_process_single_xlsx(self):
        """Test processing a single Excel file"""
        try:
            import openpyxl

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name

            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "Invoice No"
            ws['B1'] = "INV123"
            wb.save(tmp_path)

            # Process
            result = process_documents([tmp_path])

            assert isinstance(result, str)
            assert "Invoice No" in result
            assert "INV123" in result

            # Cleanup
            os.unlink(tmp_path)

        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_process_multiple_files(self):
        """Test processing multiple files together"""
        try:
            from reportlab.pdfgen import canvas
            import openpyxl

            # Create PDF
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_pdf:
                pdf_path = tmp_pdf.name

            c = canvas.Canvas(pdf_path)
            c.drawString(100, 750, "PDF Content")
            c.save()

            # Create Excel
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_xlsx:
                xlsx_path = tmp_xlsx.name

            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "Excel Content"
            wb.save(xlsx_path)

            # Process both
            result = process_documents([pdf_path, xlsx_path])

            assert isinstance(result, str)
            # Should contain content from both files
            assert len(result) > 0

            # Cleanup
            os.unlink(pdf_path)
            os.unlink(xlsx_path)

        except ImportError:
            pytest.skip("Required libraries not installed")

    def test_process_empty_list(self):
        """Test processing empty file list"""
        result = process_documents([])
        assert isinstance(result, str)
        assert len(result) == 0

    def test_process_nonexistent_file(self):
        """Test processing non-existent file raises error"""
        with pytest.raises(Exception):
            process_documents(["/nonexistent/file.pdf"])
