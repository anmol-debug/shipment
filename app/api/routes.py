from fastapi import APIRouter, UploadFile, File, HTTPException, Body, Query
from fastapi.responses import FileResponse
from typing import List, Dict, Any, Optional
import os
import tempfile
import shutil
import json
from pathlib import Path
from datetime import datetime

from app.services.document_processor import process_documents
from app.services.llm_service import extract_field_from_document
from app.services.audit_service import audit_service

router = APIRouter()

# Create uploads directory for storing files
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Create storage directory for saved extractions
STORAGE_DIR = Path("storage")
STORAGE_DIR.mkdir(exist_ok=True)


@router.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "ok", "message": "API is running"}


@router.post("/extract", response_model=dict)
async def extract_shipment_data(
    files: List[UploadFile] = File(...),
    use_mock: bool = True  # Default to mock mode to bypass Claude API
):
    """
    Extract shipment data from uploaded documents (PDF and XLSX).

    Args:
        files: Uploaded PDF/XLSX documents
        use_mock: If True, return sample data without calling Claude API (default: True)
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    temp_file_paths = []
    saved_files = []

    try:
        # Save uploaded files
        for file in files:
            # Validate file type
            file_ext = os.path.splitext(file.filename)[1].lower()
            if file_ext not in [".pdf", ".xlsx", ".xls"]:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid file type: {file.filename}. Only PDF and XLSX files are allowed."
                )

            # Save to temp file for processing
            temp_file = tempfile.NamedTemporaryFile(suffix=file_ext, delete=False)
            temp_file_paths.append(temp_file.name)

            content = await file.read()
            temp_file.write(content)
            temp_file.close()

            # Also save to uploads directory for frontend access
            saved_path = UPLOAD_DIR / file.filename
            with open(saved_path, "wb") as f:
                f.write(content)

            saved_files.append({
                "originalName": file.filename,
                "path": f"/uploads/{file.filename}",
                "size": len(content)
            })

        # MOCK MODE: Return sample data matching seed.sql structure
        if use_mock:
            print(f"[MOCK MODE] Returning sample shipment data for {len(files)} files")
            extracted_data = {
                # Core extraction fields (used in UI forms)
                "billOfLadingNumber": "ZMLU34110002",
                "containerNumber": "MSCU1234567",
                "consigneeName": "KABOFER TRADING INC",
                "consigneeAddress": "3838 CAMINO DEL RIO NORTH, STE 235, SAN DIEGO, CA 92108",
                "dateOfExport": "08/22/2019",
                "lineItemsCount": 18,
                "averageGrossWeight": "902.78 KG",
                "averagePrice": "$1289.51",

                # Additional fields from seed.sql schema
                "notes": "MOCK DATA - Sample shipment for testing audit/history features",
                "units": "CTNS",
                "containers": [{
                    "seal_number": "EMCWDGDSD",
                    "container_size": "40HQ",
                    "container_number": "MSCU1234567",
                    "container_weight": "16250",
                    "container_quantity": "776",
                    "container_measurement": "136",
                    "container_weight_unit": "KGS",
                    "container_quantity_unit": "CARTON(S)",
                    "container_measurement_unit": "CBM"
                }],
                "vessel_name": "COSCO BELGIUM",
                "voyage_number": "095E",
                "shipper_name": "CHINA ABRASIVES EXPORT CORPORATION",
                "consignee_name": "KABOFER TRADING INC",
                "port_of_loading": "SHANGHAI,CHINA",
                "port_of_discharge": "LONG BEACH, CA",
                "house_bol_number": "ZMLU34110002",
                "master_bol_number": "COSU534343282",
                "description_of_goods": "HOUSEHOLD AND PERSONAL ITEMS",
                "date_of_export_mmddyy": "082219",
                "estimated_arrival_date_mmddyy": "082919"
            }

            return {
                "success": True,
                "data": extracted_data,
                "files": saved_files,
                "mock_mode": True
            }

        # REAL MODE: Process documents and call Claude API
        # Process documents - extract text
        print(f"Processing {len(temp_file_paths)} documents...")
        document_text = process_documents(temp_file_paths)

        # Extract structured data using LLM
        print("Extracting data with Claude AI...")
        extracted_data = extract_field_from_document(document_text)

        return {
            "success": True,
            "data": extracted_data,
            "files": saved_files,
            "mock_mode": False
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error processing documents: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process documents: {str(e)}")

    finally:
        # Clean up temp files
        for path in temp_file_paths:
            try:
                if os.path.exists(path):
                    os.unlink(path)
            except Exception as e:
                print(f"Error deleting temp file {path}: {e}")


@router.get("/uploads/{filename}")
async def get_uploaded_file(filename: str):
    """
    Serve uploaded files for the frontend to display.
    """
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(file_path)


@router.get("/excel-preview/{filename}")
async def get_excel_preview(filename: str):
    """
    Convert Excel file to HTML table for preview.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        file_path = UPLOAD_DIR / filename
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Build HTML with all sheets
        html = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background: #f5f5f5;
                }
                .sheet-tabs {
                    display: flex;
                    gap: 5px;
                    margin-bottom: 10px;
                    border-bottom: 2px solid #ddd;
                }
                .sheet-tab {
                    padding: 8px 16px;
                    background: #e0e0e0;
                    border: 1px solid #ccc;
                    border-bottom: none;
                    cursor: pointer;
                    border-radius: 5px 5px 0 0;
                    font-weight: 500;
                }
                .sheet-tab.active {
                    background: white;
                    border-bottom: 2px solid white;
                    margin-bottom: -2px;
                }
                .sheet-content {
                    display: none;
                    background: white;
                    padding: 20px;
                    border: 1px solid #ddd;
                    border-radius: 0 5px 5px 5px;
                    overflow: auto;
                }
                .sheet-content.active {
                    display: block;
                }
                table {
                    border-collapse: collapse;
                    width: 100%;
                    font-size: 13px;
                }
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px 12px;
                    text-align: left;
                    white-space: nowrap;
                }
                th {
                    background-color: #4CAF50;
                    color: white;
                    font-weight: bold;
                    position: sticky;
                    top: 0;
                    z-index: 10;
                }
                tr:nth-child(even) { background-color: #f9f9f9; }
                tr:hover { background-color: #f0f0f0; }
                .empty-cell { color: #999; font-style: italic; }
                h2 {
                    color: #333;
                    margin-top: 0;
                    font-size: 18px;
                }
            </style>
            <script>
                function showSheet(sheetIndex) {
                    // Hide all sheets
                    var contents = document.getElementsByClassName('sheet-content');
                    for (var i = 0; i < contents.length; i++) {
                        contents[i].classList.remove('active');
                    }
                    // Remove active from all tabs
                    var tabs = document.getElementsByClassName('sheet-tab');
                    for (var i = 0; i < tabs.length; i++) {
                        tabs[i].classList.remove('active');
                    }
                    // Show selected sheet
                    document.getElementById('sheet-' + sheetIndex).classList.add('active');
                    document.getElementById('tab-' + sheetIndex).classList.add('active');
                }
            </script>
        </head>
        <body>
        """

        # Add sheet tabs if multiple sheets
        if len(workbook.sheetnames) > 1:
            html += '<div class="sheet-tabs">'
            for idx, sheet_name in enumerate(workbook.sheetnames):
                active = 'active' if idx == 0 else ''
                html += f'<div class="sheet-tab {active}" id="tab-{idx}" onclick="showSheet({idx})">{sheet_name}</div>'
            html += '</div>'

        # Add sheet contents
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            active = 'active' if sheet_idx == 0 else ''

            html += f'<div class="sheet-content {active}" id="sheet-{sheet_idx}">'
            html += f'<h2>{sheet_name}</h2>'
            html += '<table>'

            # Get all rows
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                # First row as header
                html += '<tr>'
                for cell in rows[0]:
                    cell_value = str(cell) if cell is not None else ''
                    html += f'<th>{cell_value}</th>'
                html += '</tr>'

                # Data rows
                for row in rows[1:]:
                    html += '<tr>'
                    for cell in row:
                        if cell is None or str(cell).strip() == '':
                            html += '<td class="empty-cell">-</td>'
                        else:
                            cell_value = str(cell)
                            html += f'<td>{cell_value}</td>'
                    html += '</tr>'

            html += '</table></div>'

        html += '</body></html>'

        from fastapi.responses import HTMLResponse
        return HTMLResponse(content=html)

    except Exception as e:
        print(f"Error generating Excel preview: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate preview: {str(e)}")


@router.post("/save")
async def save_extraction(data: Dict[str, Any] = Body(...)):
    """
    Save extracted data and file references organized by B/L number.
    """
    try:
        extracted_data = data.get("extractedData", {})
        files = data.get("files", [])

        # Use B/L number as the folder name
        bl_number = extracted_data.get("billOfLadingNumber", "unknown")
        if not bl_number or bl_number == "unknown":
            raise HTTPException(status_code=400, detail="B/L number is required to save data")

        # Sanitize B/L number for folder name
        safe_bl_number = "".join(c if c.isalnum() else "_" for c in bl_number)

        # Create folder for this B/L
        bl_folder = STORAGE_DIR / safe_bl_number
        bl_folder.mkdir(exist_ok=True)

        # Copy files to storage folder
        stored_files = []
        for file_info in files:
            original_name = file_info.get("originalName", "")
            source_path = UPLOAD_DIR / original_name

            if source_path.exists():
                dest_path = bl_folder / original_name
                shutil.copy2(source_path, dest_path)
                stored_files.append({
                    "originalName": original_name,
                    "path": str(dest_path.relative_to(STORAGE_DIR))
                })

        # Save metadata
        metadata = {
            "billOfLadingNumber": bl_number,
            "savedAt": datetime.now().isoformat(),
            "extractedData": extracted_data,
            "files": stored_files
        }

        metadata_path = bl_folder / "metadata.json"
        with open(metadata_path, "w") as f:
            json.dump(metadata, f, indent=2)

        return {
            "success": True,
            "message": f"Data saved successfully for B/L {bl_number}",
            "storagePath": str(bl_folder.relative_to(STORAGE_DIR))
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error saving data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to save data: {str(e)}")


@router.get("/retrieve/{bl_number}")
async def retrieve_extraction(bl_number: str):
    """
    Retrieve saved extraction data by B/L number.
    """
    try:
        # Sanitize B/L number
        safe_bl_number = "".join(c if c.isalnum() else "_" for c in bl_number)

        bl_folder = STORAGE_DIR / safe_bl_number
        metadata_path = bl_folder / "metadata.json"

        if not metadata_path.exists():
            raise HTTPException(status_code=404, detail=f"No saved data found for B/L {bl_number}")

        with open(metadata_path, "r") as f:
            metadata = json.load(f)

        return {
            "success": True,
            "data": metadata
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error retrieving data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve data: {str(e)}")


@router.get("/list")
async def list_saved_extractions():
    """
    List all saved B/L numbers.
    """
    try:
        saved_bls = []

        for bl_folder in STORAGE_DIR.iterdir():
            if bl_folder.is_dir():
                metadata_path = bl_folder / "metadata.json"
                if metadata_path.exists():
                    with open(metadata_path, "r") as f:
                        metadata = json.load(f)
                        saved_bls.append({
                            "billOfLadingNumber": metadata.get("billOfLadingNumber"),
                            "savedAt": metadata.get("savedAt"),
                            "containerNumber": metadata.get("extractedData", {}).get("containerNumber"),
                            "consigneeName": metadata.get("extractedData", {}).get("consigneeName")
                        })

        # Sort by saved date (most recent first)
        saved_bls.sort(key=lambda x: x.get("savedAt", ""), reverse=True)

        return {
            "success": True,
            "count": len(saved_bls),
            "data": saved_bls
        }

    except Exception as e:
        print(f"Error listing saved data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to list saved data: {str(e)}")


@router.get("/shipments/user/{user_id}")
async def get_user_shipments(user_id: str):
    """
    Get all shipments for a specific user from database

    Args:
        user_id: UUID of the user

    Returns:
        List of shipments with id, title, status, extracted_data, created_at, updated_at
    """
    try:
        from app.services.supabase_client import get_supabase

        supabase = get_supabase()

        # Query shipments for this user
        result = supabase.table('shipment_requests')\
            .select('id, title, description, status, extracted_data, created_at, updated_at, transportMode')\
            .eq('user_id', user_id)\
            .order('created_at', desc=True)\
            .execute()

        return {
            "success": True,
            "user_id": user_id,
            "count": len(result.data),
            "shipments": result.data
        }

    except Exception as e:
        print(f"Error getting user shipments: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get user shipments: {str(e)}")


# ============================================================================
# AUDIT & VERSIONING ENDPOINTS
# ============================================================================

@router.get("/shipments/{shipment_id}/history")
async def get_shipment_history(
    shipment_id: str,
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0)
):
    """
    Get audit history for a shipment

    Returns a list of all changes made to the shipment with:
    - Version numbers
    - Who made the change
    - When it was made
    - What changed
    """
    try:
        history = await audit_service.get_shipment_history(
            shipment_id=shipment_id,
            limit=limit,
            offset=offset
        )

        return {
            "success": True,
            "shipment_id": shipment_id,
            "count": len(history),
            "history": history
        }

    except Exception as e:
        print(f"Error getting shipment history: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/shipments/{shipment_id}/versions/{version_no}")
async def get_shipment_version(
    shipment_id: str,
    version_no: int
):
    """
    Get a specific version of a shipment

    Returns the complete shipment data as it existed at that version
    """
    try:
        version = await audit_service.get_version(
            shipment_id=shipment_id,
            version_no=version_no
        )

        if not version:
            raise HTTPException(
                status_code=404,
                detail=f"Version {version_no} not found for shipment {shipment_id}"
            )

        return {
            "success": True,
            "version": version
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting version: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/shipments/{shipment_id}/restore")
async def restore_shipment_version(
    shipment_id: str,
    data: Dict[str, Any] = Body(...)
):
    """
    Restore a shipment to a previous version

    Required body:
    - source_version_no: Version number to restore from
    - actor_id: User performing the restore
    - actor_name: Display name of user
    - reason: Why restoring this version
    """
    try:
        source_version_no = data.get("source_version_no")
        actor_id = data.get("actor_id")
        actor_name = data.get("actor_name")
        reason = data.get("reason", "")

        if not all([source_version_no, actor_id, actor_name]):
            raise HTTPException(
                status_code=400,
                detail="Missing required fields: source_version_no, actor_id, actor_name"
            )

        result = await audit_service.restore_version(
            shipment_id=shipment_id,
            source_version_no=source_version_no,
            actor_id=actor_id,
            actor_name=actor_name,
            reason=reason
        )

        return result

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error restoring version: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/shipments/{shipment_id}/audit")
async def create_audit_event(
    shipment_id: str,
    data: Dict[str, Any] = Body(...)
):
    """
    Create an audit event for a shipment change

    Required body:
    - event_type: Type of change (created, updated, status_changed, file_added, file_removed, restored)
    - actor_id: User making the change
    - actor_name: Display name of user
    - field_changes: What changed (dict)
    - snapshot_data: Complete current state (dict)
    - reason: Optional reason for change
    - metadata: Optional additional context
    """
    try:
        event_type = data.get("event_type")
        actor_id = data.get("actor_id")
        actor_name = data.get("actor_name")
        field_changes = data.get("field_changes", {})
        snapshot_data = data.get("snapshot_data", {})
        reason = data.get("reason")
        metadata = data.get("metadata")

        if not all([event_type, actor_id, actor_name, snapshot_data]):
            raise HTTPException(
                status_code=400,
                detail="Missing required fields: event_type, actor_id, actor_name, snapshot_data"
            )

        result = await audit_service.create_audit_event(
            shipment_id=shipment_id,
            event_type=event_type,
            actor_id=actor_id,
            actor_name=actor_name,
            reason=reason,
            field_changes=field_changes,
            snapshot_data=snapshot_data,
            metadata=metadata
        )

        return result

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error creating audit event: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/shipments/{shipment_id}/filter")
async def filter_audit_events(
    shipment_id: str,
    actor_id: Optional[str] = Query(None),
    event_type: Optional[str] = Query(None),
    field_name: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=100)
):
    """
    Filter audit events by various criteria

    Query parameters:
    - actor_id: Filter by who made the change
    - event_type: Filter by type of change
    - field_name: Filter by which field was changed
    - start_date: Filter events after this date (ISO format)
    - end_date: Filter events before this date (ISO format)
    - limit: Max number of results
    """
    try:
        # Parse dates if provided
        start_dt = datetime.fromisoformat(start_date) if start_date else None
        end_dt = datetime.fromisoformat(end_date) if end_date else None

        events = await audit_service.filter_events(
            shipment_id=shipment_id,
            actor_id=actor_id,
            event_type=event_type,
            start_date=start_dt,
            end_date=end_dt,
            field_name=field_name,
            limit=limit
        )

        return {
            "success": True,
            "shipment_id": shipment_id,
            "count": len(events),
            "events": events,
            "filters": {
                "actor_id": actor_id,
                "event_type": event_type,
                "field_name": field_name,
                "start_date": start_date,
                "end_date": end_date
            }
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        print(f"Error filtering events: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
