import openpyxl

def extract_text_from_xlsx(file_path: str) -> str:
    """
    Extract text from an XLSX file.

    Args:
        file_path: Path to the XLSX file

    Returns:
        str: Extracted text from all sheets in the XLSX file
    """
    text = ""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n\n=== Sheet: {sheet_name} ===\n"

            for row in sheet.iter_rows(values_only=True):
                # Filter out empty cells and join with |
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(cell for cell in row_data):  # Only add non-empty rows
                    text += " | ".join(row_data) + "\n"

        workbook.close()
        return text
    except Exception as e:
        raise Exception(f"Failed to extract XLSX text: {str(e)}")
